import openpyxl

# 워크북 인스턴스 객체 => excel 파일 생성
wb=openpyxl.Workbook()
# 활성화된 워크북에 워크시트 객체 -> 시트 만들기
sheet=wb.active
# 워크시트 제목 -> 제목
sheet.title='회원정보'
# 헤더컬럼
header_titles=['아이디','전화번호']
for idx, title in enumerate(header_titles):
    sheet.cell(row=1, column=idx+1,value=title)  #컬럼 idx는 0부터 시작해서 +1해줌

# 내용 저장
members=[('happy','010-1232-1231'),('smile','010-9874-2342')]

row_num=2    #1번줄은 header title이라서 2부터 시작

for r, member in enumerate(members):
    for c,v in enumerate(member):
        sheet.cell(row=row_num, column=c+1, value=v)
    row_num=row_num+1

wb.save('members.xlsx')
wb.close()